# -*- coding:utf-8 -*-
"""
2022.07.19 cuilei@idss-cn.com; 拉取Prometheus接口指标，并做相应计算，输出巡检指标并写入excel。
"""
from prometheus_api_client import PrometheusConnect
import time
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import *
import itertools
import numpy as np
import logging
import re
import requests
import json

logging.basicConfig(level=logging.INFO,  # 控制台打印的日志级别
                    filename='server_check.log',
                    filemode='a',  ##模式，有w和a，w就是写模式，每次都会重新写日志，覆盖之前的日志 a是追加模式，默认如果不写的话，就是追加模式
                    format='%(asctime)s - %(levelname)s: %(message)s'
                    )


class PrometheusApi:

    # def __init__(self, url, header):
    #     self.url = url
    #     self.header = header
    #     self.prom_client = PrometheusConnect(url=self.url, headers=self.header, disable_ssl=True)

    def __init__(self, url, headers):
        self.url = url
        self.headers = headers
        self.prom_client = PrometheusConnect(url=self.url, disable_ssl=True, headers=self.headers)

    def query_info_static(self, querystr, keywords, valuekey, replace=''):
        """
        # 取返回值中metric段中信息
        :param querystr:  PromQL查询语句
        :param keywords:  返回字典的key所取的key值
        :param valuekey:  取值的value所用的key
        :param replace:  替换字符串
        :return:  {'ip': 'value'}
        """
        result = {}
        res = self.prom_client.custom_query(query=querystr)
        for item in res:
            key = item['metric'][keywords].replace(replace, '')
            value = item['metric'][valuekey]
            result[key] = value

        return result

    def query_value_int(self, querystr, keywords, valuepos=1, replace=''):
        """
        # 取返回值中 value段中信息 ，数据为int类型
        :param querystr:  PromQL查询语句
        :param keywords:  返回字典的key所取的key值
        :param valuepos: 取值的value列表中的序列 ，通常为1
        :param replace:  替换字符串
        :return:  {'ip': 'value'}
        """
        result = {}
        res = self.prom_client.custom_query(query=querystr)
        for item in res:
            key = item['metric'][keywords].replace(replace, '')
            value = item['value'][valuepos]
            if value == " ":
                value = 'N/A'
            result[key] = value

        return result

    def query_value_percent(self, querystr, keywords, valuepos=1, replace='', ispercent='Y'):
        """
        # 取返回值中 value段中信息 ，数据为int类型
        :param querystr:  PromQL查询语句
        :param keywords:  返回字典的key所取的key值
        :param valuepos: 取值的value列表中的序列 ，通常为1
        :param replace:  替换字符串
        :return:  {'ip': 'value'}
        """
        result = {}
        res = self.prom_client.custom_query(query=querystr)
        for item in res:
            key = item['metric'][keywords].replace(replace, '')
            value = item['value'][valuepos]
            if ispercent == 'Y':
                value = str('%.2f' % float(value)) + '%'
            else:
                value = str('%.2f' % float(value))
            result[key] = value

        return result

    def query_kafka_consumer(self, querystr):
        """
        :param querystr: promQL查询语句
        :return: (consumergroup,topic,partition,sum,date)
        # kafka消费组积压指标查询专用
        """
        res = self.prom_client.custom_query(query=querystr)

        # 取出消费组名称、Topic、topic分区数、消息积压数量、时间，返回列表
        consumergroup = [i['metric']['consumergroup'] for i in res]
        topic = [i['metric']['topic'] for i in res]
        partition = [i['metric']['partition'] for i in res]
        sum = [i['value'][1] for i in res]
        date = [time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(i['value'][0])) for i in res]
        result = (sorted(zip(consumergroup, topic, partition, sum, date)))

        return result


def s_to_d(seconds):
    days = str("%.1f" % (float(seconds) // (24 * 60 * 60))) + '天'
    return days


def ms_to_d(mil_seconds):
    days = str("%.1f" % (float(mil_seconds) // (24 * 60 * 60 * 1000))) + '天'
    return days


def bytes_to_mb(rbytes):
    mbs = str("%.2f" % (float(rbytes) // (1024 * 1024))) + 'MB'
    return mbs


def color_returns(val, low=0, high=0):
    # 判断是cpu 内存 磁盘占比还是磁盘、网络读写
    # 颜色默认值设置： #FFFFFF
    color = '#FFFFFF'  # black
    try:
        if high > 0:
            if '%' in val:
                mark = re.findall(r'\d+\.?\d*', val)
                if float(mark[0]) >= float(high):  # 测试用了20.0 ,实际用80。0
                    color = '#FF7621'  # light red
                else:
                    pass
            elif 'MBps' in val:  #
                mark = re.findall(r'\d+\.?\d*', val)
                if float(mark[0]) >= float(high):
                    color = '#FF7621'  # light red
                else:
                    pass
            elif 'MB' in val:  #
                mark = re.findall(r'\d+\.?\d*', val)
                if float(mark[0]) >= float(high):
                    color = '#FF7621'  # light red
                else:
                    pass
            else:
                if float(val) >= float(high):
                    color = '#FF7621'  # light red
                else:
                    pass
        elif low > 0:
            if '%' in val:
                mark = re.findall(r'\d+\.?\d*', val)
                if float(mark[0]) < float(low):
                    color = '#FF7621'  # light red
                else:
                    pass
            elif 'MBps' in val:  #
                mark = re.findall(r'\d+\.?\d*', val)
                if float(mark[0]) < float(low):
                    color = '#FF7621'  # light red
                else:
                    pass
            else:
                if float(val) < float(low):
                    color = '#FF7621'  # light red
                else:
                    pass
    except ValueError:
        pass

    return f'background-color: {color}'


def color_returns_str(val, stype=False):
    # 判断是cpu 内存 磁盘占比还是磁盘、网络读写
    # 颜色默认值设置： #FFFFFF
    color = '#FFFFFF'  # black
    # print(val)
    if stype == 'es':
        if val == 'green':
            color = '#B8E986'  # light grenn
        elif val == 'yellow':
            color = '#F8E71C'
        else:
            color = '#FF7621'
    else:
        if val == '正常':
            color = '#B8E986'  # light grenn
        else:
            color = '#FF7621'

    return f'background-color: {color}'


def export_es_excel(cluster, node, filename):
    """
    :param cluster:  集群数据
    :param node: 节点数据
    :param filename:  文件名
    :return:
    """
    # 生成集群报表
    try:
        # 将字典列表转为DataFrame
        cluster_name = cluster[0]['cluster']
        cluster_pf = pd.DataFrame(list(cluster))
        node_pf = pd.DataFrame(list(node))
        # 设定集群属性字段顺序
        cluster_order = [
            'cluster',
            'healthy',
            'pending_task',
            'total_shards',
            'primary_shards',
            'init_shards',
            'unassign_shards',
            'delayed_shards',
            'relocating_shards',
            'breaker_tripped',
            'qps'
        ]
        cluster_pf = cluster_pf[cluster_order]
        cluster_columns_map = {
            'cluster': '集群名称',
            'healthy': '健康状态',
            'pending_task': '任务积压',
            'total_shards': '总分片数',
            'primary_shards': '主分片数',
            'init_shards': '初始化分片数',
            'unassign_shards': '未分配分片数',
            'delayed_shards': '延迟分配分片数',
            'relocating_shards': '重分配分片数',
            'breaker_tripped': '熔断器触发次数',
            'qps': '集群QPS'
        }
        cluster_pf.rename(columns=cluster_columns_map, inplace=True)
        # 设定节点属性字段顺序
        node_order = [
            'es_node_name',
            'es_node_ip',
            'es_node_roles',
            'es_node_cpu',
            'es_node_heap',
            'es_node_disk',
            'es_node_fd',
            'es_node_ygc',
            'es_node_ygc_time',
            'es_node_ogc',
            'es_node_ogc_time',
            'es_node_bulk_reject',
            'es_node_search_reject',
            'es_node_shards',
            'es_node_qps',
            'es_node_iops',
            'es_node_indices_size',
            'es_node_indices_docs'
        ]
        node_pf = node_pf[node_order]
        node_columns_map = {
            'es_node_name': '节点名称',
            'es_node_ip': '节点IP',
            'es_node_roles': '节点角色',
            'es_node_cpu': '节点ES进程CPU使用率',
            'es_node_heap': '节点堆内存使用率',
            'es_node_disk': 'ES数据盘磁盘使用率',
            'es_node_fd': 'ES进程文件描述符使用率',
            'es_node_ygc': '节点youngGC次数',
            'es_node_ygc_time': '节点youngGC耗时',
            'es_node_ogc': '节点oldGC次数',
            'es_node_ogc_time': '节点oldGC耗时',
            'es_node_bulk_reject': '节点批量写入失败次数',
            'es_node_search_reject': '节点搜索失败次数',
            'es_node_shards': '节点分片个数',
            'es_node_qps': '节点QPS',
            'es_node_iops': '节点IOPS',
            'es_node_indices_size': 'data节点数据量',
            'es_node_indices_docs': 'data节点总文档数量'
        }
        node_pf.rename(columns=node_columns_map, inplace=True)
        # 设定写入文件句柄
        filewriter = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        # 写入集群状态
        cluster_pf.fillna(' ', inplace=True)
        cluster_pf.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='ES巡检报告')

        # 写入节点状态
        node_pf.fillna(' ', inplace=True)
        cluster_rows = cluster_pf.shape[0]
        node_pf.to_excel(filewriter, sheet_name='ES巡检报告', startrow=cluster_rows+3, index=False)
        #  设置表格格式，自动列宽
        column_width = (
            node_pf.columns.to_series().apply(lambda x: len(x.encode('gbk'))).values
        )
        max_width = (
            node_pf.astype(str).applymap(lambda x: len(x.encode('gbk'))).agg(max).values
        )
        widths = np.max((column_width, max_width), axis=0)
        worksheet = filewriter.sheets['ES巡检报告']
        for i, width in enumerate(widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width + 2

        # 设置表格格式：添加边框
        border_set = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for cell in itertools.chain(*worksheet):
            if cell.value is None:
                pass
            else:
                cell.border = border_set
        #
        # 集群指标着色
        cluster_pf0 = cluster_pf.style.applymap(color_returns_str, stype='es', subset=['健康状态'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='ES巡检报告')
        cluster_pf0 = cluster_pf.style.applymap(color_returns, high=1, subset=['任务积压'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='ES巡检报告')
        # 节点指标着色 使用率
        node_pf0 = node_pf.style.applymap(color_returns, high=80, subset=[
            '节点ES进程CPU使用率', '节点堆内存使用率', 'ES数据盘磁盘使用率', 'ES进程文件描述符使用率'
        ])
        node_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='ES巡检报告', startrow=cluster_rows+3)
        # 节点指标着色 GC
        node_pf0 = node_pf.style.applymap(color_returns, high=1000, subset=[
            '节点youngGC次数', '节点youngGC耗时', '节点oldGC次数', '节点oldGC耗时'
        ])
        node_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='ES巡检报告', startrow=cluster_rows+3)
        # 节点指标着色 thread_pool失败
        node_pf0 = node_pf.style.applymap(color_returns, high=1, subset=[
            '节点批量写入失败次数', '节点搜索失败次数'
        ])
        node_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='ES巡检报告', startrow=cluster_rows+3)
        # 最终写入
        cluster_pf.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='ES巡检报告')

        #保存表格
        filewriter.save()

    except Exception as e:
        print(e)
        logging.exception(e)


def export_kafka_excel(cluster, consumer, filename):
    """
    :param cluster:  集群数据
    :param consumer: 消费组数据
    :param filename:  文件名
    :return:
    """
    # 生成集群报表
    try:
        # 将字典列表转为DataFrame
        cluster_pf = pd.DataFrame(list(cluster))

        # 设定集群属性字段顺序
        cluster_order = [
            'kafka_brokers',
            'kafka_lag',
            'kafka_topic',
            'kafka_partitions',
            'kafka_replicas',
            'kafka_sync',
            'kafka_under_replicated',
            'kafka_mps'
        ]
        cluster_pf = cluster_pf[cluster_order]
        cluster_columns_map = {
            'kafka_brokers': '节点数量',
            'kafka_lag': '集群消息积压',
            'kafka_topic': 'TOPIC数量',
            'kafka_partitions': '分区数量',
            'kafka_replicas': '副本数量',
            'kafka_sync': '同步中副本数量',
            'kafka_under_replicated': '失效副本数量',
            'kafka_mps': '每秒消息吞吐量'
        }
        cluster_pf.rename(columns=cluster_columns_map, inplace=True)
        # 设定节点属性字段顺序
        consumer_order = [
            'consumergroup',
            'topic',
            'partition',
            'lag',
            'date'
        ]
        consumer_pf = pd.DataFrame(consumer, columns=consumer_order)
        consumer_columns_map = {
            'consumergroup': '消费组',
            'topic': 'TOPIC',
            'partition': '分区',
            'lag': '消息积压',
            'date': '时间'
        }
        consumer_pf.rename(columns=consumer_columns_map, inplace=True)
        # 设定写入文件句柄
        filewriter = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        # 写入集群状态
        cluster_pf.fillna(' ', inplace=True)
        cluster_pf.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='kafka巡检报告')
        # 写入节点状态
        consumer_pf.fillna(' ', inplace=True)
        cluster_rows = cluster_pf.shape[0]
        consumer_pf.to_excel(filewriter, sheet_name='kafka巡检报告', startrow=cluster_rows+3, index=False)
        #  设置表格格式，自动列宽
        column_width = (
            cluster_pf.columns.to_series().apply(lambda x: len(x.encode('gbk'))).values
        )
        max_width = (
            cluster_pf.astype(str).applymap(lambda x: len(x.encode('gbk'))).agg(max).values
        )
        widths = np.max((column_width, max_width), axis=0)
        worksheet = filewriter.sheets['kafka巡检报告']
        for i, width in enumerate(widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width + 2

        # 设置表格格式：添加边框
        border_set = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for cell in itertools.chain(*worksheet):
            if cell.value is None:
                pass
            else:
                cell.border = border_set
        #
        # 集群指标着色
        cluster_pf0 = cluster_pf.style.applymap(color_returns, low=3, subset=['节点数量'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='kafka巡检报告')
        cluster_pf0 = cluster_pf.style.applymap(color_returns, high=5000, subset=['集群消息积压'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='kafka巡检报告')
        cluster_pf0 = cluster_pf.style.applymap(color_returns, high=1, subset=['失效副本数量'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='kafka巡检报告')
        # 消费组指标着色 消息积压
        consumer_pf0 = consumer_pf.style.applymap(color_returns, high=2000, subset=['消息积压'])
        consumer_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='kafka巡检报告', startrow=cluster_rows + 3)

        cluster_pf.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='kafka巡检报告')

        # 保存表格
        filewriter.save()

    except Exception as e:
        print(e)
        logging.exception(e)


def export_zk_excel(cluster, filename):
    """
    :param cluster:  集群数据
    :param filename:  文件名
    :return:
    """
    # 生成集群报表
    try:
        # 将字典列表转为DataFrame
        cluster_pf = pd.DataFrame(list(cluster))
        # 设定集群属性字段顺序
        cluster_order = [
            'zk_ip',
            'zk_up',
            'zk_fd',
            'zk_fd_used',
            'zk_lag_request',
            'zk_pending',
            'zk_latency',
            'zk_send_packets'
        ]
        cluster_pf = cluster_pf[cluster_order]
        cluster_columns_map = {
            'zk_ip': 'IP地址',
            'zk_up': '状态',
            'zk_fd': '文件句柄比例',
            'zk_fd_used': '已用文件句柄',
            'zk_lag_request': '堆积请求数',
            'zk_pending': '阻塞sync数',
            'zk_latency': '响应延时',
            'zk_send_packets': '发送数据包量'
        }
        cluster_pf.rename(columns=cluster_columns_map, inplace=True)
        # 设定写入文件句柄
        filewriter = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        # 写入集群状态
        cluster_pf.fillna(' ', inplace=True)
        cluster_pf.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='zookeeper巡检报告')
        #  设置表格格式，自动列宽
        column_width = (
            cluster_pf.columns.to_series().apply(lambda x: len(x.encode('gbk'))).values
        )
        max_width = (
            cluster_pf.astype(str).applymap(lambda x: len(x.encode('gbk'))).agg(max).values
        )
        widths = np.max((column_width, max_width), axis=0)
        worksheet = filewriter.sheets['zookeeper巡检报告']
        for i, width in enumerate(widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width + 2

        # 设置表格格式：添加边框
        border_set = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for cell in itertools.chain(*worksheet):
            if cell.value is None:
                pass
            else:
                cell.border = border_set
        #
        cluster_pf.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='zookeeper巡检报告')
        # 集群指标着色
        cluster_pf0 = cluster_pf.style.applymap(color_returns_str, stype='other', subset=['状态'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='zookeeper巡检报告')
        cluster_pf0 = cluster_pf.style.applymap(color_returns, high=80, subset=['文件句柄比例'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='zookeeper巡检报告')
        cluster_pf0 = cluster_pf.style.applymap(color_returns, high=100, subset=['堆积请求数', '阻塞sync数'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='zookeeper巡检报告')
        cluster_pf0 = cluster_pf.style.applymap(color_returns, high=1000, subset=['响应延时'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='zookeeper巡检报告')

        # 最终写入
        cluster_pf.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='zookeeper巡检报告')

        # 保存表格
        filewriter.save()

    except Exception as e:
        print(e)
        logging.exception(e)


def export_my_excel(cluster, filename):
    """
    :param cluster:  集群数据
    :param filename:  文件名
    :return:
    """
    # 生成集群报表
    try:
        # 将字典列表转为DataFrame
        cluster_pf = pd.DataFrame(list(cluster))
        # 设定集群属性字段顺序
        cluster_order = [
            'my_ip',
            'my_status',
            'my_up',
            'my_qps',
            'my_conn',
            'my_fd',
            'my_fd_used',
            'my_slow',
            'my_tl'
        ]
        cluster_pf = cluster_pf[cluster_order]
        cluster_columns_map = {
            'my_ip': 'IP地址',
            'my_status': '运行状态',
            'my_up': '运行时间',
            'my_qps': 'QPS',
            'my_conn': '连接数占比',
            'my_fd': '已用文件句柄占比',
            'my_fd_used': '已用文件句柄',
            'my_slow': '最近24小时慢查询',
            'my_tl': '近5分钟TableLock'
        }
        cluster_pf.rename(columns=cluster_columns_map, inplace=True)
        # 设定写入文件句柄
        filewriter = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        # 写入集群状态
        cluster_pf.fillna(' ', inplace=True)
        cluster_pf.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='mysql巡检报告')
        #  设置表格格式，自动列宽
        column_width = (
            cluster_pf.columns.to_series().apply(lambda x: len(x.encode('gbk'))).values
        )
        max_width = (
            cluster_pf.astype(str).applymap(lambda x: len(x.encode('gbk'))).agg(max).values
        )
        widths = np.max((column_width, max_width), axis=0)
        worksheet = filewriter.sheets['mysql巡检报告']
        for i, width in enumerate(widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width + 2

        # 设置表格格式：添加边框
        border_set = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for cell in itertools.chain(*worksheet):
            if cell.value is None:
                pass
            else:
                cell.border = border_set
        #
        # 指标着色
        cluster_pf0 = cluster_pf.style.applymap(color_returns_str, stype='other', subset=['运行状态'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='mysql巡检报告')
        cluster_pf0 = cluster_pf.style.applymap(color_returns, high=80, subset=['连接数占比', '已用文件句柄占比'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='mysql巡检报告')
        cluster_pf0 = cluster_pf.style.applymap(color_returns, high=500, subset=['最近24小时慢查询'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='mysql巡检报告')
        # 最终写入
        cluster_pf.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='mysql巡检报告')

        # 保存表格
        filewriter.save()

    except Exception as e:
        print(e)
        logging.exception(e)


def export_redis_excel(cluster, filename):
    """
    :param cluster:  集群数据
    :param filename:  文件名
    :return:
    """
    # 生成集群报表
    try:
        # 将字典列表转为DataFrame
        cluster_pf = pd.DataFrame(list(cluster))
        # 设定集群属性字段顺序
        cluster_order = [
            'redis_ip',
            'redis_status',
            'redis_up',
            'redis_conn',
            'redis_cps',
            'redis_hit',
            'redis_miss',
            'redis_mem_use',
            'redis_input',
            'redis_output'
        ]
        cluster_pf = cluster_pf[cluster_order]
        cluster_columns_map = {
            'redis_ip': 'IP地址',
            'redis_status': '运行状态',
            'redis_up': '运行时间',
            'redis_conn': '当前连接数',
            'redis_cps': '每秒命令执行',
            'redis_hit': '缓存命中率',
            'redis_miss': '缓存miss率',
            'redis_mem_use': '内存占用',
            'redis_input': '请求流量',
            'redis_output': '响应流量'
        }
        cluster_pf.rename(columns=cluster_columns_map, inplace=True)
        # 设定写入文件句柄
        filewriter = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        # 写入集群状态
        cluster_pf.fillna(' ', inplace=True)
        cluster_pf.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='redis巡检报告')
        #  设置表格格式，自动列宽
        column_width = (
            cluster_pf.columns.to_series().apply(lambda x: len(x.encode('gbk'))).values
        )
        max_width = (
            cluster_pf.astype(str).applymap(lambda x: len(x.encode('gbk'))).agg(max).values
        )
        widths = np.max((column_width, max_width), axis=0)
        worksheet = filewriter.sheets['redis巡检报告']
        for i, width in enumerate(widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width + 2

        # 设置表格格式：添加边框
        border_set = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for cell in itertools.chain(*worksheet):
            if cell.value is None:
                pass
            else:
                cell.border = border_set
        #
        # 指标着色
        cluster_pf0 = cluster_pf.style.applymap(color_returns_str, stype='other', subset=['运行状态'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='redis巡检报告')
        cluster_pf0 = cluster_pf.style.applymap(color_returns, high=1000, subset=['当前连接数'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='redis巡检报告')
        cluster_pf0 = cluster_pf.style.applymap(color_returns, low=30, subset=['缓存命中率'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='redis巡检报告')
        cluster_pf0 = cluster_pf.style.applymap(color_returns, high=60, subset=['缓存miss率'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='redis巡检报告')
        cluster_pf0 = cluster_pf.style.applymap(color_returns, high=4000, subset=['内存占用'])
        cluster_pf0.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='redis巡检报告')

        # 最终写入
        cluster_pf.to_excel(filewriter, encoding='utf-8', index=False, sheet_name='redis巡检报告')

        # 保存表格
        filewriter.save()

    except Exception as e:
        print(e)
        logging.exception(e)


def get_es_response(esapi, query):
    start_time = 1635955200000
    end_time = 1636041599000

    headers = {
        'content-type': 'application/json'
    }

    index = 'security_log_' + '2022.08.03'
    url = esapi + index + '/_search'
    req = requests.post(url, headers=headers, data=query)
    response = json.loads(req.content.decode('utf-8'))

    return response['took']